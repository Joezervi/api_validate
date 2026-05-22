"""Parse CSV and Excel files into the same product-dict format as the PDF parser.

Expected columns (header row, case-insensitive):
  SKU, Barcode, Supplier Code, Product Name, Job No., Qty, Unit Price, Subtotal

Columns are matched by name, not position.  Extra columns are ignored;
missing columns get empty strings / None.
"""

import csv
import io
import re
from typing import Any

# Canonical column names and their output keys
_COLUMN_MAP = {
    "sku":            "sku",
    "barcode":        "barcode",
    "supplier code":  "supplier_code",
    "supplier":       "supplier_code",
    "product name":   "product_name",
    "product":        "product_name",
    "job no.":        "job_no",
    "job no":         "job_no",
    "job":            "job_no",
    "qty":            "qty",
    "quantity":       "qty",
    "unit price":     "unit_price",
    "price":          "unit_price",
    "subtotal":       "subtotal",
}


def _normalize(col_name: str) -> str:
    """Strip whitespace, lower-case, collapse multiple spaces."""
    return re.sub(r"\s+", " ", (col_name or "").strip().lower())


def _parse_value(key: str, raw: str) -> Any:
    """Convert a raw cell value to the appropriate type for the given key."""
    if raw is None:
        return None
    raw = str(raw).strip()
    if not raw or raw.lower() in ("none", "null", "n/a", "-"):
        return None

    if key in ("qty",):
        try:
            return int(float(raw))
        except (ValueError, TypeError):
            return raw  # keep as string for user to fix

    if key in ("unit_price", "subtotal"):
        cleaned = re.sub(r"[^\d.\-]", "", raw.replace(",", "."))
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return raw

    return raw


def parse_csv(file_bytes: bytes) -> list[dict]:
    """Parse a CSV file (bytes) into a list of product dicts."""
    wrapper = io.TextIOWrapper(io.BytesIO(file_bytes), encoding="utf-8-sig")
    reader = csv.DictReader(wrapper)
    if not reader.fieldnames:
        return []

    # Build column index mapping from header
    col_map: dict[str, str] = {}
    for i, name in enumerate(reader.fieldnames):
        norm = _normalize(name)
        if norm in _COLUMN_MAP:
            col_map[_COLUMN_MAP[norm]] = name

    products = []
    for row in reader:
        item = {}
        for key, col_name in col_map.items():
            val = _parse_value(key, row.get(col_name, ""))
            if val is not None:
                item[key] = val

        if item.get("sku") and item.get("product_name"):
            products.append(item)

    return products


def parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse an Excel file (.xlsx / .xls) into a list of product dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel parsing")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    # Read header row
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return []

    header = [_normalize(str(c) if c is not None else "") for c in rows[0]]

    # Build column index map
    col_idx: dict[str, int] = {}
    for i, norm in enumerate(header):
        if norm in _COLUMN_MAP:
            key = _COLUMN_MAP[norm]
            if key not in col_idx:  # first match wins
                col_idx[key] = i

    products = []
    for row in rows[1:]:
        item = {}
        for key, idx in col_idx.items():
            raw = row[idx] if idx < len(row) else None
            val = _parse_value(key, raw)
            if val is not None:
                item[key] = val

        if item.get("sku") and item.get("product_name"):
            products.append(item)

    return products


def parse_spreadsheet(file_bytes: bytes, filename: str) -> list[dict]:
    """Auto-detect format and parse a CSV or Excel file.

    Parameters
    ----------
    file_bytes : bytes
        Raw file content.
    filename : str
        Original filename, used to detect format by extension.

    Returns
    -------
    list[dict]
        Each dict has keys matching the 8-column PO format.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("csv", "tsv", "txt"):
        return parse_csv(file_bytes)
    elif ext in ("xlsx", "xls", "xlsm"):
        return parse_excel(file_bytes)
    else:
        # Fallback: try CSV first, then Excel
        try:
            result = parse_csv(file_bytes)
            if result:
                return result
        except Exception:
            pass
        try:
            return parse_excel(file_bytes)
        except Exception:
            pass
        return []
